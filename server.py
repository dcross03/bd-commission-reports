from flask import Flask, request, jsonify, redirect
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import base64, io, os, json, requests

app = Flask(__name__)
CORS(app)

TEMPLATE_B64 = os.environ.get('TEMPLATE_B64', '')
CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '')
CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
REFRESH_TOKEN = os.environ.get('GOOGLE_REFRESH_TOKEN', '')
REDIRECT_URI = 'https://bd-commission-reports.onrender.com/oauth/callback'
SCOPES = 'https://www.googleapis.com/auth/gmail.compose'

def get_access_token():
    resp = requests.post('https://oauth2.googleapis.com/token', data={
        'client_id': CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'refresh_token': REFRESH_TOKEN,
        'grant_type': 'refresh_token',
    })
    return resp.json().get('access_token')

def parse_date(s):
    if not s: return None
    for fmt in ('%m/%d/%Y','%m/%d/%y','%Y-%m-%dT%H:%M:%S','%Y-%m-%d %H:%M:%S','%Y-%m-%d'):
        try: return datetime.strptime(str(s).strip(), fmt)
        except: pass
    return None

def build_report(email, cc, month_label, orders):
    wb = load_workbook(io.BytesIO(base64.b64decode(TEMPLATE_B64)))
    ws = wb.active
    for row_num in range(11, ws.max_row + 1):
        for col in range(1, 9):
            ws.cell(row_num, col).value = None
    ws['G2'] = month_label
    ws['H4'] = f'=SUM(E11:E{10+len(orders)})'
    ws['H7'] = email
    if cc:
        ws['G8'] = 'CC:'
        ws['G8'].font = Font(name='Calibri', size=11, bold=True)
        ws['G8'].alignment = Alignment(horizontal='center')
        ws['H8'] = cc
        ws['H8'].font = Font(name='Calibri', size=11)
    else:
        ws['G8'] = None
        ws['H8'] = None
    CF = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
    for i, o in enumerate(orders):
        r = 11 + i
        def sc(col, value, numfmt=None, halign=None):
            cell = ws.cell(r, col)
            cell.value = value
            cell.font = Font(name='Calibri', size=11)
            if halign: cell.alignment = Alignment(horizontal=halign)
            if numfmt: cell.number_format = numfmt
        pd = parse_date(o.get('poDate',''))
        dpd = parse_date(o.get('datePaid',''))
        sc(2, pd or o.get('poDate',''), 'mm-dd-yy', 'left')
        sc(3, str(o.get('po','')), None, 'left')
        sc(4, o.get('cust',''), None, 'center')
        sc(5, float(o.get('total',0)), CF, 'center')
        sc(6, dpd or o.get('datePaid',''), 'mm-dd-yy', 'center')
        sc(7, o.get('rg',''), None, 'center')
        sc(8, o.get('rn',''), None, 'center')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def create_draft(access_token, to, cc, subject, body_text, attachment_bytes, attachment_filename):
    msg = MIMEMultipart()
    msg['To'] = to
    if cc:
        msg['Cc'] = cc
    msg['Subject'] = subject
    msg.attach(MIMEText(body_text, 'plain'))

    part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    part.set_payload(attachment_bytes)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{attachment_filename}"')
    msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    resp = requests.post(
        'https://gmail.googleapis.com/gmail/v1/users/me/drafts',
        headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'},
        json={'message': {'raw': raw}}
    )
    return resp.status_code == 200, resp.json()

@app.route('/build', methods=['POST'])
def build():
    data = request.json
    results = {}
    for rep in data.get('reports', []):
        xlsx_bytes = build_report(rep['email'], rep.get('cc'), rep['monthLabel'], rep['orders'])
        results[rep['group']] = base64.b64encode(xlsx_bytes).decode()
    return jsonify(results)

@app.route('/send-drafts', methods=['POST'])
def send_drafts():
    if not REFRESH_TOKEN:
        return jsonify({'error': 'Not authorized. Visit /authorize first.'}), 401

    access_token = get_access_token()
    if not access_token:
        return jsonify({'error': 'Could not get access token. Re-authorize at /authorize.'}), 401

    data = request.json
    results = []
    mm = data.get('mm', '')
    yy = data.get('yy', '')

    for rep in data.get('reports', []):
        group = rep['group']
        to = rep['email']
        cc = rep.get('cc')
        subject = f"{mm}/{yy} Commission Report"
        body = f"Hello,\n\nAttached is your commission report for {rep['monthLabel']}. If you have any questions, please let us know."
        filename = f"{group.replace(' ','_').replace('/','_')}_Commission_Report_{mm}-{yy}.xlsx"

        xlsx_bytes = build_report(to, cc, rep['monthLabel'], rep['orders'])
        ok, resp = create_draft(access_token, to, cc, subject, body, xlsx_bytes, filename)
        results.append({'group': group, 'status': 'ok' if ok else 'error', 'detail': str(resp)})

    return jsonify({'results': results})

@app.route('/authorize')
def authorize():
    url = (
        'https://accounts.google.com/o/oauth2/v2/auth'
        f'?client_id={CLIENT_ID}'
        f'&redirect_uri={REDIRECT_URI}'
        f'&response_type=code'
        f'&scope={SCOPES}'
        '&access_type=offline'
        '&prompt=consent'
    )
    return redirect(url)

@app.route('/oauth/callback')
def oauth_callback():
    code = request.args.get('code')
    if not code:
        return 'Error: no code received.', 400
    resp = requests.post('https://oauth2.googleapis.com/token', data={
        'code': code,
        'client_id': CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'redirect_uri': REDIRECT_URI,
        'grant_type': 'authorization_code',
    })
    tokens = resp.json()
    refresh_token = tokens.get('refresh_token', '')
    if refresh_token:
        return (
            f'<h2>Authorization successful!</h2>'
            f'<p>Copy this refresh token and add it to Render as <strong>GOOGLE_REFRESH_TOKEN</strong>:</p>'
            f'<p style="word-break:break-all;background:#f0f0f0;padding:12px;font-family:monospace">{refresh_token}</p>'
            f'<p>Once added, redeploy your Render service and the email drafting will work.</p>'
        )
    return f'Error getting refresh token: {tokens}', 400

@app.route('/')
def health():
    return jsonify({'status': 'ok', 'authorized': bool(REFRESH_TOKEN)})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
